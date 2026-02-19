#!/usr/bin/env python3

# Contest Management System - http://cms-dev.github.io/
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

"""Excel export utilities and handlers for Training Programs.

This module contains Excel formatting utilities and export handlers for
attendance and combined ranking data.

"""

import io
import json
import logging
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from cms.db import TrainingDay, TrainingProgram, Student
from .analysis import (
    PairInfo,
    apply_location_weights,
    apply_training_type_correction,
    calculate_time_decay_weights,
    calculate_weighted_averages,
    collect_student_td_info,
    drop_outliers,
    get_raw_scores,
    normalize_scores,
    run_pairwise_analysis,
)
from .base import BaseHandler, require_permission
from .training_analytics import (
    TrainingProgramFilterMixin,
    get_attendance_view_data,
    get_ranking_view_data,
    FilterContext,
)

logger = logging.getLogger(__name__)


# ----------------------------------------------------------------------------
# Styles & Constants
# ----------------------------------------------------------------------------

SLUG_REGEX = r"[^A-Za-z0-9_-]+"

EXCEL_ZEBRA_COLORS = [
    ("4472C4", "D9E2F3"),
    ("70AD47", "E2EFDA"),
    ("ED7D31", "FCE4D6"),
    ("7030A0", "E4DFEC"),
    ("00B0F0", "DAEEF3"),
    ("FFC000", "FFF2CC"),
]

STYLE_HEADER_FONT = Font(bold=True)
STYLE_HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
STYLE_BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
STYLE_FILL_BLUE = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
STYLE_FILL_GREY = PatternFill(
    start_color="808080", end_color="808080", fill_type="solid"
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


# ----------------------------------------------------------------------------
# Utilities
# ----------------------------------------------------------------------------


def excel_safe(value: Any) -> Any:
    """Escape potentially dangerous Excel values."""
    if isinstance(value, str) and value and re.match(r"^\s*[=+\-@]", value):
        return "'" + value
    return value


def build_filename(program_name: str, export_type: str, ctx: FilterContext) -> str:
    """Build a filename for Excel export based on context."""
    slug = re.sub(SLUG_REGEX, "_", program_name)
    parts = [slug, export_type]

    if ctx.start_date:
        parts.append(f"from_{ctx.start_date.strftime('%Y%m%d')}")
    if ctx.end_date:
        parts.append(f"to_{ctx.end_date.strftime('%Y%m%d')}")
    if ctx.training_day_types:
        t_slug = re.sub(SLUG_REGEX, "_", "_".join(ctx.training_day_types))
        parts.append(f"types_{t_slug}")
    if ctx.student_tags:
        t_slug = re.sub(SLUG_REGEX, "_", "_".join(ctx.student_tags))
        parts.append(f"tags_{t_slug}")

    return "_".join(parts) + ".xlsx"


class TrainingExcelWriter:
    """Helper to manage worksheet writing and styling."""

    def __init__(self, ws: Worksheet):
        self.ws = ws

    def setup_static_headers(self):
        """Write the Student and Tags columns."""
        for col, title in enumerate(["Student", "Tags"], start=1):
            cell = self.ws.cell(row=1, column=col, value=title)
            cell.font = STYLE_HEADER_FONT_WHITE
            cell.fill = STYLE_FILL_BLUE
            cell.border = STYLE_BORDER_THIN
            cell.alignment = ALIGN_CENTER
            self.ws.merge_cells(
                start_row=1, start_column=col, end_row=2, end_column=col
            )

    def write_student_meta(self, row: int, student: Student):
        """Write student name and tags."""
        if student.participation:
            u = student.participation.user
            name = f"{u.first_name} {u.last_name} ({u.username})"
        else:
            name = "(Unknown)"

        # Name
        self.ws.cell(
            row=row, column=1, value=excel_safe(name)
        ).border = STYLE_BORDER_THIN
        # Tags
        tags = "; ".join(student.student_tags) if student.student_tags else ""
        self.ws.cell(
            row=row, column=2, value=excel_safe(tags)
        ).border = STYLE_BORDER_THIN

    def write_named_header(
        self, col: int, title: str, width: int, idx: int
    ) -> tuple[int, PatternFill]:
        """Write a zebra-colored merged header row and return (next_col, sub_fill)."""
        h_color, sub_color = EXCEL_ZEBRA_COLORS[idx % len(EXCEL_ZEBRA_COLORS)]
        fill = PatternFill(start_color=h_color, end_color=h_color, fill_type="solid")
        sub_fill = PatternFill(
            start_color=sub_color, end_color=sub_color, fill_type="solid"
        )

        cell = self.ws.cell(row=1, column=col, value=excel_safe(title))
        cell.font = STYLE_HEADER_FONT_WHITE
        cell.fill = fill
        cell.border = STYLE_BORDER_THIN
        cell.alignment = ALIGN_CENTER

        self.ws.merge_cells(
            start_row=1, start_column=col, end_row=1, end_column=col + width - 1
        )
        return col + width, sub_fill

    def write_td_header(
        self, col: int, td: Any, width: int, idx: int
    ) -> tuple[int, PatternFill]:
        """Write a merged Training Day header and return the next column index."""
        return self.write_named_header(col, _td_title(td), width, idx)

    def write_subheaders(self, col: int, headers: list[str], fill: PatternFill):
        """Write the second row of headers."""
        for i, text in enumerate(headers):
            cell = self.ws.cell(row=2, column=col + i, value=text)
            cell.font = STYLE_HEADER_FONT
            cell.fill = fill
            cell.border = STYLE_BORDER_THIN
            cell.alignment = Alignment(horizontal="center")

    def auto_size_columns(self, max_col: int):
        """Apply basic widths."""
        self.ws.column_dimensions["A"].width = 30
        self.ws.column_dimensions["B"].width = 20
        for i in range(3, max_col + 1):
            self.ws.column_dimensions[get_column_letter(i)].width = 12


# ----------------------------------------------------------------------------
# Sheet Generators
# ----------------------------------------------------------------------------

def generate_attendance_sheet(ws: Worksheet, view_data: dict, context: FilterContext):
    """Populate the worksheet with attendance data."""
    writer = TrainingExcelWriter(ws)
    writer.setup_static_headers()

    subcols = ["Status", "Location", "Recorded", "Delay Reasons", "Comments"]
    width = len(subcols)

    # Write Headers
    curr_col = 3
    for i, td in enumerate(context.archived_training_days):
        _, sub_fill = writer.write_td_header(curr_col, td, width, i)
        writer.write_subheaders(curr_col, subcols, sub_fill)
        curr_col += width

    # Write Rows
    row = 3
    attendance_map = view_data["attendance_data"]

    for student in view_data["sorted_students"]:
        writer.write_student_meta(row, student)

        curr_col = 3
        for td in context.archived_training_days:
            att = attendance_map.get(student.id, {}).get(td.id)

            if att:
                # Status Logic
                if att.status == "missed":
                    status = "Justified Absent" if att.justified else "Missed"
                elif att.delay_time:
                    mins = att.delay_time.total_seconds() / 60
                    status = (
                        f"Delayed ({mins:.0f}m)"
                        if mins < 60
                        else f"Delayed ({mins / 60:.1f}h)"
                    )
                else:
                    status = "On Time"

                # Location Logic
                loc_map = {"class": "Class", "home": "Home", "both": "Both"}
                loc = (
                    loc_map.get(att.location, att.location)
                    if att.status != "missed"
                    else ""
                )

                if att.status == "missed":
                    rec = ""
                else:
                    rec = "Yes" if att.recorded else "No"

                vals = [status, loc, rec, att.delay_reasons or "", att.comment or ""]
            else:
                vals = ["", "", "", "", ""]

            for i, val in enumerate(vals):
                cell = ws.cell(row=row, column=curr_col + i, value=excel_safe(val))
                cell.border = STYLE_BORDER_THIN

            curr_col += width
        row += 1

    writer.auto_size_columns(curr_col)


def generate_ranking_sheet(ws: Worksheet, view_data: dict):
    """Populate the worksheet with ranking data."""
    writer = TrainingExcelWriter(ws)
    writer.setup_static_headers()

    td_list = view_data["filtered_training_days"]
    tasks_map = view_data["training_day_tasks"]
    ranking_map = view_data["ranking_data"]

    # Write Headers
    curr_col = 3
    for i, td in enumerate(td_list):
        tasks = tasks_map.get(td.id, [])
        width = len(tasks) + 1  # +1 for Total

        _, sub_fill = writer.write_td_header(curr_col, td, width, i)

        subheaders = [t["name"] for t in tasks] + ["Total"]
        writer.write_subheaders(curr_col, subheaders, sub_fill)

        curr_col += width

    # Global Total Header
    ws.cell(row=1, column=curr_col, value="Global").fill = STYLE_FILL_GREY
    ws.cell(row=1, column=curr_col).font = STYLE_HEADER_FONT_WHITE
    ws.cell(row=1, column=curr_col).border = STYLE_BORDER_THIN
    ws.cell(row=1, column=curr_col).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=1, start_column=curr_col, end_row=2, end_column=curr_col)

    # Write Rows
    row = 3
    for student in view_data["sorted_students"]:
        writer.write_student_meta(row, student)

        curr_col = 3
        global_total = 0.0

        for td in td_list:
            tasks = tasks_map.get(td.id, [])
            rank = ranking_map.get(student.id, {}).get(td.id)
            td_total = 0.0

            # Task Scores
            for task in tasks:
                val = None
                if rank and rank.task_scores:
                    val = rank.task_scores.get(str(task["id"]))

                cell = ws.cell(row=row, column=curr_col)
                if val is not None:
                    cell.value = val
                    td_total += val
                cell.border = STYLE_BORDER_THIN
                curr_col += 1

            # TD Total
            cell = ws.cell(row=row, column=curr_col)
            if rank and rank.task_scores:
                cell.value = td_total
                global_total += td_total
            cell.border = STYLE_BORDER_THIN
            curr_col += 1

        # Global Total
        cell = ws.cell(row=row, column=curr_col)
        cell.value = global_total if global_total > 0 else ""
        cell.border = STYLE_BORDER_THIN
        cell.font = STYLE_HEADER_FONT

        row += 1

    writer.auto_size_columns(curr_col)


# ----------------------------------------------------------------------------
# Handlers
# ----------------------------------------------------------------------------

class ExportAttendanceHandler(TrainingProgramFilterMixin, BaseHandler):
    """Export attendance data to Excel format."""

    @require_permission(BaseHandler.AUTHENTICATED)
    def get(self, training_program_id: str):
        tp = self.safe_get_item(TrainingProgram, training_program_id)
        ctx = self.get_filter_context(tp)

        if not ctx.archived_training_days:
            self.redirect(self.url("training_program", tp.id, "attendance"))
            return

        view_data = get_attendance_view_data(ctx)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        generate_attendance_sheet(ws, view_data, ctx)

        self._serve_excel(wb, build_filename(tp.name, "attendance", ctx))

    def _serve_excel(self, wb: Workbook, filename: str):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        self.set_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.set_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.write(output.getvalue())
        self.finish()


class ExportCombinedRankingHandler(ExportAttendanceHandler):
    """Export combined ranking data to Excel format."""

    @require_permission(BaseHandler.AUTHENTICATED)
    def get(self, training_program_id: str):
        tp = self.safe_get_item(TrainingProgram, training_program_id)
        ctx = self.get_filter_context(tp)

        # Build Data
        view_data = get_ranking_view_data(ctx)

        if not view_data["filtered_training_days"]:
            self.redirect(self.url("training_program", tp.id, "combined_ranking"))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Combined Ranking"

        generate_ranking_sheet(ws, view_data)

        self._serve_excel(wb, build_filename(tp.name, "ranking", ctx))


def _write_weights_grid(
    ws: Worksheet,
    sorted_students: list[Student],
    columns: list[tuple[int, str]],
    weights_data: dict[int, dict[int, float]],
):
    """Shared logic for writing a weights sheet (regular or paired)."""
    writer = TrainingExcelWriter(ws)
    writer.setup_static_headers()

    curr_col = 3
    for i, (_, title) in enumerate(columns):
        _, sub_fill = writer.write_named_header(curr_col, title, 1, i)
        writer.write_subheaders(curr_col, ["Weight"], sub_fill)
        curr_col += 1

    row = 3
    for student in sorted_students:
        writer.write_student_meta(row, student)
        curr_col = 3
        for col_id, _ in columns:
            w = weights_data.get(student.id, {}).get(col_id)
            cell = ws.cell(row=row, column=curr_col)
            if w is not None:
                cell.value = round(w, 4)
                cell.number_format = '0.0000'
            cell.border = STYLE_BORDER_THIN
            curr_col += 1
        row += 1

    writer.auto_size_columns(curr_col)


def _write_scores_grid(
    ws: Worksheet,
    sorted_students: list[Student],
    columns: list[tuple[int, str]],
    scores_data: dict[int, dict[int, float]],
    weighted_averages: dict[int, float],
    method: str,
):
    """Shared logic for writing a normalized-scores sheet (regular or paired)."""
    writer = TrainingExcelWriter(ws)
    writer.setup_static_headers()

    label = "Rank" if method == "rank" else "Score"
    curr_col = 3
    for i, (_, title) in enumerate(columns):
        _, sub_fill = writer.write_named_header(curr_col, title, 1, i)
        writer.write_subheaders(curr_col, [label], sub_fill)
        curr_col += 1

    avg_label = "Weighted Avg Rank" if method == "rank" else "Weighted Avg"
    cell = ws.cell(row=1, column=curr_col, value=avg_label)
    cell.fill = STYLE_FILL_GREY
    cell.font = STYLE_HEADER_FONT_WHITE
    cell.border = STYLE_BORDER_THIN
    cell.alignment = ALIGN_CENTER
    ws.merge_cells(
        start_row=1, start_column=curr_col, end_row=2, end_column=curr_col
    )

    fmt = '0' if method == "rank" else '0.00'

    row = 3
    for student in sorted_students:
        writer.write_student_meta(row, student)
        curr_col = 3
        for col_id, _ in columns:
            val = scores_data.get(student.id, {}).get(col_id)
            cell = ws.cell(row=row, column=curr_col)
            if val is not None:
                cell.value = round(val, 4) if method != "rank" else val
                cell.number_format = fmt
            cell.border = STYLE_BORDER_THIN
            curr_col += 1

        avg = weighted_averages.get(student.id)
        cell = ws.cell(row=row, column=curr_col)
        if avg is not None:
            cell.value = round(avg, 4) if method != "rank" else round(avg, 2)
            cell.number_format = '0.00'
        cell.border = STYLE_BORDER_THIN
        cell.font = STYLE_HEADER_FONT
        row += 1

    writer.auto_size_columns(curr_col)


def _td_title(td: TrainingDay) -> str:
    """Build a human-readable header for a single training day."""
    title = td.description or td.name or "Session"
    if td.start_time:
        title += f" ({td.start_time.strftime('%b %d')})"
    if td.training_day_types:
        title += f" [{'; '.join(td.training_day_types)}]"
    return title


def generate_weights_sheet(
    ws: Worksheet,
    sorted_students: list[Student],
    training_days: list[TrainingDay],
    student_weights: dict[int, dict[int, float]],
):
    """Populate a worksheet with per-student per-training-day weights."""
    columns = [(td.id, _td_title(td)) for td in training_days]
    _write_weights_grid(ws, sorted_students, columns, student_weights)


def generate_normalized_scores_sheet(
    ws: Worksheet,
    sorted_students: list[Student],
    training_days: list[TrainingDay],
    normalized_scores: dict[int, dict[int, float]],
    weighted_averages: dict[int, float],
    method: str,
):
    """Populate a worksheet with normalized scores and weighted averages."""
    columns = [(td.id, _td_title(td)) for td in training_days]
    _write_scores_grid(
        ws, sorted_students, columns, normalized_scores, weighted_averages, method,
    )


def generate_paired_weights_sheet(
    ws: Worksheet,
    sorted_students: list[Student],
    pairs: list[PairInfo],
    training_days: list[TrainingDay],
    paired_weights: dict[int, dict[int, float]],
):
    """Populate a worksheet with per-student per-pair weights."""
    td_map = {td.id: td for td in training_days}
    columns = [
        (p.pair_id, _pair_title(td_map.get(p.td_a_id), td_map.get(p.td_b_id)))
        for p in pairs
    ]
    _write_weights_grid(ws, sorted_students, columns, paired_weights)


def generate_paired_scores_sheet(
    ws: Worksheet,
    sorted_students: list[Student],
    pairs: list[PairInfo],
    training_days: list[TrainingDay],
    paired_norm_scores: dict[int, dict[int, float]],
    paired_weighted_avgs: dict[int, float],
    method: str,
):
    """Populate a worksheet with paired normalized scores and weighted averages."""
    td_map = {td.id: td for td in training_days}
    columns = [
        (p.pair_id, _pair_title(td_map.get(p.td_a_id), td_map.get(p.td_b_id)))
        for p in pairs
    ]
    _write_scores_grid(
        ws, sorted_students, columns, paired_norm_scores,
        paired_weighted_avgs, method,
    )


def _pair_title(td_a, td_b) -> str:
    """Build a human-readable header for a TD pair."""
    def _short(td):
        if td is None:
            return "?"
        name = td.description or td.name or "Session"
        if td.start_time:
            name += f" ({td.start_time.strftime('%b %d')})"
        return name
    return "{" + _short(td_a) + ", " + _short(td_b) + "}"


class ExportAnalysedRankingHandler(ExportAttendanceHandler):
    """Export analysed ranking data (weights + normalized scores) to Excel."""

    @require_permission(BaseHandler.AUTHENTICATED)
    def post(self, training_program_id: str):
        tp = self.safe_get_item(TrainingProgram, training_program_id)
        ctx = self.get_filter_context(tp)

        ranking_view = get_ranking_view_data(ctx)
        attendance_view = get_attendance_view_data(ctx)

        td_list: list[TrainingDay] = ranking_view["filtered_training_days"]
        if not td_list:
            self.redirect(self.url("training_program", tp.id, "combined_ranking"))
            return

        student_info = collect_student_td_info(
            ranking_view["ranking_data"],
            attendance_view["attendance_data"],
            ranking_view["training_day_tasks"],
            td_list,
        )

        first_weight_pct = float(self.get_argument("first_training_weight", "100"))
        home_factor = float(self.get_argument("home_factor", "1.0"))
        recorded_home_factor = float(
            self.get_argument("recorded_home_factor", "1.0")
        )

        type_pcts_raw = self.get_argument("type_percentages", "{}")
        try:
            type_percentages: dict[str, float] = json.loads(type_pcts_raw)
            type_percentages = {
                k: v / 100.0 for k, v in type_percentages.items()
            }
        except (json.JSONDecodeError, ValueError, TypeError):
            type_percentages = {}

        type_assigns_raw = self.get_argument("type_assignments", "{}")
        try:
            type_assignments: dict[str, str] = json.loads(type_assigns_raw)
            type_assignments = {int(k): v for k, v in type_assignments.items()}
        except (json.JSONDecodeError, ValueError):
            type_assignments = {}

        norm_method = self.get_argument("normalization_method", "none")
        top_x = max(1, int(self.get_argument("top_x", "10")))
        normalize_variability = self.get_argument(
            "normalize_variability", "off"
        ) == "on"
        num_outliers = int(self.get_argument("num_outliers", "0"))

        base_weights = calculate_time_decay_weights(td_list, first_weight_pct)
        student_weights = apply_location_weights(
            base_weights, student_info, home_factor, recorded_home_factor
        )
        student_weights = apply_training_type_correction(
            student_weights, td_list, type_assignments, type_percentages
        )

        if num_outliers > 0:
            raw_scores = get_raw_scores(student_info, td_list)
            student_weights = drop_outliers(
                student_weights, raw_scores, num_outliers
            )

        norm_scores = normalize_scores(
            norm_method, student_info, td_list, top_x, normalize_variability
        )
        weighted_avgs = calculate_weighted_averages(student_weights, norm_scores)

        sorted_students = ranking_view["sorted_students"]

        wb = Workbook()

        ws_weights = wb.active
        ws_weights.title = "Weights"
        generate_weights_sheet(ws_weights, sorted_students, td_list, student_weights)

        ws_scores = wb.create_sheet("Normalized Scores")
        generate_normalized_scores_sheet(
            ws_scores, sorted_students, td_list,
            norm_scores, weighted_avgs, norm_method,
        )

        include_pairwise = self.get_argument("include_pairwise", "off") == "on"
        if include_pairwise and len(td_list) >= 2:
            pairs, pw, pn, pa = run_pairwise_analysis(
                student_info, student_weights, td_list,
                norm_method, top_x, normalize_variability,
            )
            if pairs:
                ws_pw = wb.create_sheet("Paired Weights")
                generate_paired_weights_sheet(
                    ws_pw, sorted_students, pairs, td_list, pw,
                )
                ws_pn = wb.create_sheet("Paired Normalized Scores")
                generate_paired_scores_sheet(
                    ws_pn, sorted_students, pairs, td_list, pn, pa, norm_method,
                )

        self._serve_excel(wb, build_filename(tp.name, "analysed", ctx))
